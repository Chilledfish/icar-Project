import re
import requests
import pandas as pd
from dataframe_to_dictionary import model_converter, normalize_datagov_models
from ncap import NCAP_compiler
from urllib.parse import quote, unquote
from openpyxl import load_workbook
import numpy as np
import xlrd
from openpyxl import load_workbook



wb = load_workbook('ncap_models.xlsx')
names = wb.sheetnames


makes = ["Aiways", "Alfa-Romeo", "Audi", "BMW", "Bentley", "Chevrolet", "Citroen", "Dacia", "DS", "DS", "Fiat",
         "Ford", "Honda", "Hyundai", "Isuzu", "Jaguar", "Jeep", "Kia", "Lancia", "Land-Rover",
         "Land-Rover", "Lexus", "Lynk & Co", "Maserati", "Mazda", "Mercedes", "MG", "Mitsubishi",
         "Nissan", "Opel", "Peugeot", "Porsche", "Renault", "Saab", "Seat",
         "Skoda", "Smart", "Ssangyong", "Subaru", "Suzuki", "Tesla", "Toyota", "Volkswagen",
         "Volvo"]

data_makes = ["איוויס", "אלפא רומיאו", "אאודי", "ב מ וו", "שברולט", "סיטרואן", "דאצ'יה", "די.אס", "פיאט", "פורד",
              "הונדה", "יונדאי", "איסוזו", "יגואר", "ג'יפ", "קיה", "לנצ'יה", "לנדרובר", "לקסוס", "לינק אנד קו",
              "מזראטי", "מזדה", "מרצדס", "מ.ג", "מיצובישי", "ניסאן", "אופל", "פיג'ו", "פורשה", "רנו", "סיאט",
              "סקודה", "סמארט", "סאנגיונג", "סובראו", "סוזוקי", "טסלה", "טויוטה", "פולקסווגן", "וולוו"]

ncap_data_dict = dict(zip(makes, data_makes))

small_makes = [make.lower() for make in makes]

make_exceptions = {"Opel/Vauxhall": "Opel",
                   "Lynk-&amp;-co": "Lynk & Co",
                   "Mini": "BMW",
                   "Genesis": "Hyundai",
                   "Cupra": "Seat",
                   "škoda": "Skoda",
                   "citroën": "Citroen",
                   "vw": "Volkswagen",
                   "mercedes-benz": "Mercedes",
                   "range-rover": "Land-Rover",
                   "infiniti": "Nissan",
                   "geely-emgrand": "Geely"
                   }

keys = make_exceptions.keys()
keylist = list(keys)

makes_dict = dict(zip(small_makes, makes))

makes_dict.update(make_exceptions)
ncap_df = pd.read_csv('ncap_safety.csv', encoding='utf-8')

data_df = pd.read_csv('datagov_safety.csv', encoding='utf-8')

data_df = data_df[data_df['shnat_yitzur'] > 2008]

#tozarim = data_df['tozar'].to_list()
#tozarim_set = set(tozarim)
#tozarim_set = list(tozarim_set)
#a = [car_model for car_model in tozarim_set if car_model == float('nan')]
#b = tozarim_set[0]
#cursor = np.isnan(b)
#d = data_df.applu

#tozarim = data_df[data_df['tozar'] == float('nan')]
#tozarim = data_df['tozar'].apply(lambda car_model: 1 if np.isnan(car_model) == True else 2)

#tozarim_nan = tozarim[tozarim['tozar'] == 1]
c = 0

with open('ncap_urls.txt', 'r', encoding='utf-8') as url_file:
	ncap_urls = [i.strip('\n') for i in url_file]



file_workbook = "ncap_models.xlsx"


a = []




	   # open an Excel file and return a workbook


#tozar, kinuy_mishari
def normalize_datagov_to_ncap(model, df_stack):
	model_column = df_stack[df_stack == model]
	model_list = model_column.index.to_list()
	a = len(model_list)
	if a > 0:
		return model_list[0][1]
	return 'del'


def sim(y):
	x = 1
	y = x + 1
	return y
make = ''
data_df2 = pd.DataFrame(columns=data_df.columns)
#moo = data_df.applymap(normalize_datagov_to_ncap)
data_df = data_df.reset_index(drop=True)



#------------------------------------------------
for i in range(data_df.shape[0]):
	row = data_df.iloc[i]
	row['kinuy_mishari'] = str(row['kinuy_mishari'].replace('-', ' ').replace('?', '').lower())
	new_make = row['tozar']
	if isinstance(new_make, float):
		continue
	if make != new_make:
		make = new_make
		#wb = load_workbook(file_workbook, read_only=True)
		if make in wb.sheetnames:
			excel_df = pd.read_excel(file_workbook, sheet_name=make, engine='openpyxl', dtype='str')
			a = excel_df.shape[0]
			excel_df.loc[excel_df.shape[0] + 1] = excel_df.columns
			excel_df.iloc[-1] = excel_df.columns
			df_stack = excel_df.stack().astype('str')
			df_stack = df_stack.apply(lambda x: str(x).replace('-', ' ').replace('?', '').lower())
	a = 1

	model = row['kinuy_mishari']
	kinuy = normalize_datagov_to_ncap(model, df_stack)
	if kinuy == 'del':
		continue
	#print(i)
	row['kinuy_mishari'] = kinuy
	data_df2.loc[i] = row
	b = 5
#--------------------------------------------------



data_df2 = data_df2.reset_index(drop=True)
data_df2.to_csv('n_datagov_safety.csv', index=None, encoding='utf-8')